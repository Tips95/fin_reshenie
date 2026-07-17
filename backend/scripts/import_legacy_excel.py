"""
Импорт legacy-клиентов из Excel «Новый УЧЕТ.xlsm».

Лист «Банкротство»  → этап bankruptcy + график + платежи.
Лист «СБОР ДОКУМЕНТОВ» → этап document_collection (только если клиента нет на банкротстве).

Использование (из каталога backend):
  python scripts/import_legacy_excel.py --excel "C:\\path\\Новый УЧЕТ.xlsm"
  python scripts/import_legacy_excel.py --excel ... --execute
  DATABASE_URL=postgresql://... python scripts/import_legacy_excel.py --excel ... --execute

Без --execute — только dry-run (отчёт, без записи в БД).
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import defaultdict
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from uuid import UUID

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker

BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))

from app.core.config import settings  # noqa: E402
from app.core.database import SessionLocal, build_connect_args, sanitize_database_url  # noqa: E402
from app.models.client import Client  # noqa: E402
from app.models.document_collection import DocumentCollection  # noqa: E402
from app.models.enums import (  # noqa: E402
    ClientStatus,
    DocumentCollectionStatus,
    EngagementStage,
    MandatoryPaymentStatus,
    MandatoryPaymentType,
    OrganizationType,
    PaymentScheduleStatus,
    ProcedureStage,
    UserRole,
)
from app.models.installment_plan import InstallmentPlan  # noqa: E402
from app.models.organization import Organization  # noqa: E402
from app.models.payment import Payment  # noqa: E402
from app.models.payment_schedule import PaymentSchedule  # noqa: E402
from app.models.user import User  # noqa: E402
from app.services.mandatory_payments import create_default_mandatory_payments  # noqa: E402

LEGACY_COLLECTION_CUTOFF = date(2026, 6, 1)
STRICT_DEBT_CUTOFF = date(2026, 7, 17)
IMPORT_TAG = "legacy_excel_import"


def mask_database_url(url: str) -> str:
    return re.sub(r"://([^:/@]+):([^@]+)@", r"://\\1:***@", url)


def is_remote_database(url: str) -> bool:
    lowered = url.lower()
    return lowered.startswith("postgresql") and "localhost" not in lowered and "127.0.0.1" not in lowered


@contextmanager
def open_import_session(database_url: str | None):
    if database_url:
        clean_url = sanitize_database_url(database_url)
        engine = create_engine(
            clean_url,
            pool_pre_ping=True,
            connect_args=build_connect_args(database_url),
        )
        session_factory = sessionmaker(bind=engine, autocommit=False, autoflush=False)
        db = session_factory()
        try:
            yield db
        finally:
            db.close()
            engine.dispose()
    else:
        db = SessionLocal()
        try:
            yield db
        finally:
            db.close()


def print_preflight(db: Session, *, database_url: str, execute: bool) -> None:
    organization, owner = get_bankruptcy_context(db)
    total_clients = db.scalar(
        select(func.count())
        .select_from(Client)
        .where(
            Client.organization_id == organization.id,
            Client.is_deleted.is_(False),
        )
    )
    collection_clients = db.scalar(
        select(func.count())
        .select_from(Client)
        .where(
            Client.organization_id == organization.id,
            Client.is_deleted.is_(False),
            Client.engagement_stage == EngagementStage.DOCUMENT_COLLECTION,
        )
    )
    bankruptcy_clients = db.scalar(
        select(func.count())
        .select_from(Client)
        .where(
            Client.organization_id == organization.id,
            Client.is_deleted.is_(False),
            Client.engagement_stage == EngagementStage.BANKRUPTCY,
        )
    )
    remote = is_remote_database(database_url)

    print("=== PREFLIGHT ===")
    print(f"Database: {mask_database_url(database_url)}")
    print(f"Remote/production-like: {'yes' if remote else 'no'}")
    print(f"Organization: {organization.name} ({organization.id})")
    print(f"Owner: {owner.email}")
    print(f"Existing clients: total={total_clients}, collection={collection_clients}, bankruptcy={bankruptcy_clients}")
    print(f"Mode: {'EXECUTE (write)' if execute else 'DRY-RUN (no writes)'}")
    print("Safety: only INSERT new clients; duplicates skipped by phone/name; no updates/deletes")
    print("=================")


def money(value: float | int | Decimal | str | None) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.strip().replace(" ", "").replace(",", ".")
        if not cleaned:
            return None
        try:
            return Decimal(cleaned).quantize(Decimal("0.01"))
        except Exception:
            return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except Exception:
        return None


def parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def norm_phone(value: object) -> str | None:
    if value is None:
        return None
    raw = str(int(value)) if isinstance(value, float) and value == int(value) else str(value)
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) != 11:
        return None
    return f"+{digits}"


def norm_name(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def parse_bankruptcy_sheet(wb) -> list[dict]:
    ws = wb["Банкротство"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    month_cols: list[tuple[int, date]] = []
    for index, cell in enumerate(header):
        parsed = parse_date(cell)
        if parsed is not None:
            month_cols.append((index, parsed))

    clients: list[dict] = []
    for row in rows[1:]:
        if row[0] is None or row[1] is None:
            continue
        name = str(row[1]).strip()
        if not name:
            continue
        monthly: list[tuple[date, Decimal]] = []
        for col_index, month_date in month_cols:
            amount = money(row[col_index] if col_index < len(row) else None)
            if amount and amount > 0:
                monthly.append((month_date, amount))
        clients.append(
            {
                "name": name,
                "phone": norm_phone(row[2]),
                "contract_date": parse_date(row[3]),
                "contract_amount": money(row[4]),
                "executor_amount": money(row[6]),
                "fin_management": money(row[7]),
                "court": money(row[8]),
                "deposit": money(row[9]),
                "profit_col": money(row[10]),
                "remainder_col": money(row[11]),
                "monthly_payments": monthly,
            }
        )
    return clients


def parse_collection_sheet(wb) -> list[dict]:
    ws = wb["СБОР ДОКУМЕНТОВ"]
    rows = list(ws.iter_rows(values_only=True))
    items: list[dict] = []
    for row in rows[1:]:
        if row[0] is None or row[1] is None:
            continue
        name = str(row[1]).strip()
        if not name:
            continue
        paid = money(row[7])
        notary = money(row[8]) or Decimal("0.00")
        expenses = money(row[6]) or Decimal("0.00")
        if paid is None:
            continue
        collection_fee = paid - notary
        if collection_fee < 0:
            collection_fee = Decimal("0.00")
        items.append(
            {
                "name": name,
                "contract_date": parse_date(row[2]),
                "phone": norm_phone(row[3]),
                "expenses": expenses,
                "paid_total": paid,
                "notary_fee": notary,
                "collection_fee": collection_fee,
                "manager_commission": Decimal("0.00"),
            }
        )
    return items


def index_collection(items: list[dict]) -> dict[str, dict]:
    by_key: dict[str, dict] = {}
    for item in items:
        if item["phone"]:
            by_key[f"phone:{item['phone']}"] = item
        by_key[f"name:{norm_name(item['name'])}"] = item
    return by_key


def lookup_collection(index: dict[str, dict], *, phone: str | None, name: str) -> dict | None:
    if phone and f"phone:{phone}" in index:
        return index[f"phone:{phone}"]
    key = f"name:{norm_name(name)}"
    return index.get(key)


def build_document_collection(
    client_id: UUID,
    *,
    paid_total: Decimal,
    collection_fee: Decimal,
    notary_fee: Decimal,
    manager_commission: Decimal,
    paid_date: date,
) -> DocumentCollection:
    return DocumentCollection(
        client_id=client_id,
        total_amount=paid_total,
        collection_fee=collection_fee,
        notary_fee=notary_fee,
        manager_commission=manager_commission,
        status=DocumentCollectionStatus.PAID,
        paid_date=paid_date,
    )


def apply_mandatory_amounts(client_id: UUID, row: dict) -> list:
    items = create_default_mandatory_payments(client_id)
    for item in items:
        if item.payment_type == MandatoryPaymentType.DEPOSIT:
            item.planned_amount = row["deposit"] or Decimal("25000.00")
        elif item.payment_type == MandatoryPaymentType.FINANCIAL_MANAGEMENT:
            item.planned_amount = row["fin_management"] or Decimal("0.00")
        elif item.payment_type == MandatoryPaymentType.COURT_FEE:
            court = row["court"] or Decimal("0.00")
            item.planned_amount = court
            item.is_applicable = court > 0
            if not item.is_applicable:
                item.status = MandatoryPaymentStatus.NOT_APPLICABLE
    return items


def build_schedule_and_payments(
    *,
    plan: InstallmentPlan,
    contract_amount: Decimal,
    contract_date: date,
    monthly_payments: list[tuple[date, Decimal]],
    remainder: Decimal | None,
    client_id: UUID,
    created_by: UUID,
) -> tuple[list[PaymentSchedule], list[Payment]]:
    schedules: list[PaymentSchedule] = []
    payments: list[Payment] = []

    for month_number, (due_date, amount) in enumerate(monthly_payments, start=1):
        schedule = PaymentSchedule(
            installment_plan_id=plan.id,
            month_number=month_number,
            due_date=due_date,
            planned_amount=amount,
            paid_amount=amount,
            paid_date=due_date,
            status=PaymentScheduleStatus.PAID,
        )
        schedules.append(schedule)

    if remainder is not None and remainder > 0:
        last_due = monthly_payments[-1][0] if monthly_payments else contract_date
        next_due = last_due + relativedelta(months=1)
        schedules.append(
            PaymentSchedule(
                installment_plan_id=plan.id,
                month_number=len(schedules) + 1,
                due_date=next_due,
                planned_amount=remainder,
                paid_amount=Decimal("0.00"),
                status=PaymentScheduleStatus.PENDING,
            )
        )

    plan.total_months = len(schedules) if schedules else max(len(monthly_payments), 1)

    return schedules, payments


def attach_payments_after_flush(
    db,
    *,
    schedules: list[PaymentSchedule],
    monthly_payments: list[tuple[date, Decimal]],
    client_id: UUID,
    created_by: UUID,
) -> list[Payment]:
    payments: list[Payment] = []
    paid_schedules = [s for s in schedules if s.status == PaymentScheduleStatus.PAID]
    for schedule, (payment_date, amount) in zip(paid_schedules, monthly_payments, strict=False):
        payments.append(
            Payment(
                client_id=client_id,
                payment_schedule_id=schedule.id,
                amount=amount,
                payment_date=payment_date,
                comment=IMPORT_TAG,
                created_by=created_by,
            )
        )
    db.add_all(payments)
    return payments


def get_bankruptcy_context(db):
    organization = db.scalar(
        select(Organization)
        .where(Organization.organization_type == OrganizationType.BANKRUPTCY)
        .limit(1)
    )
    if organization is None:
        raise RuntimeError("Организация bankruptcy не найдена в БД")

    owner = db.scalar(
        select(User).where(
            User.organization_id == organization.id,
            User.role == UserRole.OWNER,
            User.is_active.is_(True),
        )
    )
    if owner is None:
        raise RuntimeError("Owner пользователь не найден")

    return organization, owner


def existing_keys(db, organization_id: UUID) -> tuple[set[str], set[str]]:
    clients = list(
        db.scalars(
            select(Client).where(
                Client.organization_id == organization_id,
                Client.is_deleted.is_(False),
            )
        )
    )
    phones = {c.phone for c in clients if c.phone}
    names = {norm_name(c.full_name) for c in clients}
    return phones, names


def import_bankruptcy_client(
    db,
    *,
    organization_id: UUID,
    owner_id: UUID,
    row: dict,
    collection_index: dict[str, dict],
    execute: bool,
) -> dict:
    phone = row["phone"] or f"+7000{abs(hash(row['name'])) % 10_000_000:07d}"
    contract_date = row["contract_date"] or date(2025, 9, 1)
    contract_amount = row["contract_amount"]
    if contract_amount is None:
        return {"action": "skip", "reason": "нет суммы договора", "name": row["name"]}

    remainder = row["remainder_col"]
    if remainder is None:
        paid_sum = sum((amount for _, amount in row["monthly_payments"]), Decimal("0.00"))
        remainder = contract_amount - paid_sum

    collection = lookup_collection(collection_index, phone=row["phone"], name=row["name"])
    status = ClientStatus.COMPLETED if remainder <= 0 else ClientStatus.ACTIVE

    result = {
        "action": "create_bankruptcy",
        "name": row["name"],
        "phone": phone,
        "contract_date": str(contract_date),
        "contract_amount": str(contract_amount),
        "payments": len(row["monthly_payments"]),
        "remainder": str(remainder),
        "status": status.value,
        "collection_paid": str(collection["paid_total"]) if collection else None,
    }

    if not execute:
        return result

    client = Client(
        organization_id=organization_id,
        full_name=row["name"],
        phone=phone,
        contract_date=contract_date,
        debt_amount=Decimal("0.00"),
        engagement_stage=EngagementStage.BANKRUPTCY,
        procedure_stage=ProcedureStage.CONTRACT_SIGNED,
        status=status,
    )
    db.add(client)
    db.flush()

    if collection is not None:
        db.add(
            build_document_collection(
                client.id,
                paid_total=collection["paid_total"],
                collection_fee=collection["collection_fee"],
                notary_fee=collection["notary_fee"],
                manager_commission=collection["manager_commission"],
                paid_date=collection["contract_date"] or contract_date,
            )
        )
    else:
        legacy_total = Decimal("10000.00")
        db.add(
            build_document_collection(
                client.id,
                paid_total=legacy_total,
                collection_fee=legacy_total,
                notary_fee=Decimal("0.00"),
                manager_commission=Decimal("0.00"),
                paid_date=contract_date,
            )
        )

    db.add_all(apply_mandatory_amounts(client.id, row))

    plan = InstallmentPlan(
        client_id=client.id,
        pricing_tier_id=None,
        total_amount=contract_amount,
        start_date=contract_date,
        total_months=1,
    )
    db.add(plan)
    db.flush()

    schedules, _ = build_schedule_and_payments(
        plan=plan,
        contract_amount=contract_amount,
        contract_date=contract_date,
        monthly_payments=row["monthly_payments"],
        remainder=remainder,
        client_id=client.id,
        created_by=owner_id,
    )
    db.add_all(schedules)
    db.flush()
    attach_payments_after_flush(
        db,
        schedules=schedules,
        monthly_payments=row["monthly_payments"],
        client_id=client.id,
        created_by=owner_id,
    )
    return result


def import_collection_client(
    db,
    *,
    organization_id: UUID,
    row: dict,
    execute: bool,
) -> dict:
    phone = row["phone"] or f"+7100{abs(hash(row['name'])) % 10_000_000:07d}"
    contract_date = row["contract_date"] or date.today()
    result = {
        "action": "create_collection",
        "name": row["name"],
        "phone": phone,
        "paid_total": str(row["paid_total"]),
    }
    if not execute:
        return result

    client = Client(
        organization_id=organization_id,
        full_name=row["name"],
        phone=phone,
        contract_date=contract_date,
        debt_amount=Decimal("0.00"),
        engagement_stage=EngagementStage.DOCUMENT_COLLECTION,
        procedure_stage=ProcedureStage.CONTRACT_SIGNED,
        status=ClientStatus.ACTIVE,
    )
    db.add(client)
    db.flush()
    db.add(
        build_document_collection(
            client.id,
            paid_total=row["paid_total"],
            collection_fee=row["collection_fee"],
            notary_fee=row["notary_fee"],
            manager_commission=row["manager_commission"],
            paid_date=contract_date,
        )
    )
    return result


def run_import(excel_path: Path, execute: bool, *, database_url: str | None, allow_remote: bool) -> int:
    if not excel_path.is_file():
        print(f"Файл не найден: {excel_path}")
        return 1

    effective_url = database_url or os.environ.get("DATABASE_URL") or settings.DATABASE_URL
    if not effective_url:
        print("DATABASE_URL не задан")
        return 1

    if execute and is_remote_database(effective_url) and not allow_remote:
        print("Для записи в удалённую PostgreSQL добавьте флаг --allow-remote")
        return 1

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    bankruptcy_rows = parse_bankruptcy_sheet(wb)
    collection_rows = parse_collection_sheet(wb)
    collection_index = index_collection(collection_rows)
    wb.close()

    with open_import_session(database_url) as db:
        print_preflight(db, database_url=effective_url, execute=execute)
        organization, owner = get_bankruptcy_context(db)
        existing_phones, existing_names = existing_keys(db, organization.id)

        results: list[dict] = []
        bankruptcy_keys: set[str] = set()

        for row in bankruptcy_rows:
            key_phone = row["phone"]
            key_name = norm_name(row["name"])
            if key_phone and key_phone in existing_phones:
                results.append({"action": "skip", "name": row["name"], "reason": "телефон уже в базе"})
                continue
            if key_name in existing_names:
                results.append({"action": "skip", "name": row["name"], "reason": "ФИО уже в базе"})
                continue
            if key_phone:
                bankruptcy_keys.add(f"phone:{key_phone}")
            bankruptcy_keys.add(f"name:{key_name}")
            results.append(
                import_bankruptcy_client(
                    db,
                    organization_id=organization.id,
                    owner_id=owner.id,
                    row=row,
                    collection_index=collection_index,
                    execute=execute,
                )
            )
            if execute:
                if key_phone:
                    existing_phones.add(key_phone)
                existing_names.add(key_name)

        for row in collection_rows:
            phone = row["phone"]
            name_key = f"name:{norm_name(row['name'])}"
            phone_key = f"phone:{phone}" if phone else None
            if phone_key and phone_key in bankruptcy_keys:
                continue
            if name_key in bankruptcy_keys:
                continue
            if phone and phone in existing_phones:
                results.append({"action": "skip", "name": row["name"], "reason": "телефон уже в базе"})
                continue
            if norm_name(row["name"]) in existing_names:
                results.append({"action": "skip", "name": row["name"], "reason": "ФИО уже в базе"})
                continue
            results.append(
                import_collection_client(
                    db,
                    organization_id=organization.id,
                    row=row,
                    execute=execute,
                )
            )
            if execute:
                if phone:
                    existing_phones.add(phone)
                existing_names.add(norm_name(row["name"]))

        if execute:
            db.commit()
            print("IMPORT COMPLETE")
        else:
            db.rollback()
            print("DRY-RUN (no database writes)")

        counts = defaultdict(int)
        for item in results:
            counts[item["action"]] += 1

        print(f"Bankruptcy rows in Excel: {len(bankruptcy_rows)}")
        print(f"Collection rows in Excel: {len(collection_rows)}")
        print("Summary:", dict(counts))
        for item in results:
            if item.get("reason"):
                print(f"  SKIP {item['name']}: {item['reason']}")
            elif item["action"] == "create_bankruptcy":
                print(
                    f"  BANK {item['name']} | {item['contract_amount']} RUB | "
                    f"payments {item['payments']} | remainder {item['remainder']}"
                )
            elif item["action"] == "create_collection":
                print(f"  COLL {item['name']} | collection {item['paid_total']} RUB")
        return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="Импорт legacy Excel в Fin Reshenie")
    parser.add_argument(
        "--excel",
        required=True,
        help="Путь к .xlsm / .xlsx",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Записать в БД (без флага — только dry-run)",
    )
    parser.add_argument(
        "--database-url",
        default=None,
        help="Переопределить DATABASE_URL (для prod PostgreSQL Timeweb)",
    )
    parser.add_argument(
        "--allow-remote",
        action="store_true",
        help="Разрешить запись в удалённую PostgreSQL (prod)",
    )
    args = parser.parse_args()
    if args.database_url:
        os.environ["DATABASE_URL"] = args.database_url
    raise SystemExit(
        run_import(
            Path(args.excel),
            execute=args.execute,
            database_url=args.database_url,
            allow_remote=args.allow_remote,
        )
    )


if __name__ == "__main__":
    main()
