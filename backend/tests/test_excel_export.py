import uuid
from datetime import date, datetime
from decimal import Decimal
from types import SimpleNamespace
from io import BytesIO
from unittest.mock import MagicMock

from openpyxl import load_workbook

from app.models.enums import ClientStatus, EngagementStage, PaymentScheduleStatus, ProcedureStage, UserRole
from app.schemas.client import ClientDetailResponse
from app.schemas.installment_plan import InstallmentPlanResponse
from app.schemas.mandatory_payment import MandatoryPaymentResponse
from app.schemas.payment import PaymentResponse
from app.schemas.payment_schedule import PaymentScheduleResponse
from app.services.excel_export import (
    build_client_detail_workbook,
    build_clients_workbook,
    build_overdue_clients_workbook,
    client_overdue_amount,
    workbook_to_bytes,
)


ORG_ID = uuid.UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")
USER_ID = uuid.UUID("bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb")
CLIENT_ID = uuid.UUID("cccccccc-cccc-cccc-cccc-cccccccccccc")
MANAGER_ID = uuid.UUID("dddddddd-dddd-dddd-dddd-dddddddddddd")


def make_user(role: UserRole = UserRole.OWNER) -> SimpleNamespace:
    return SimpleNamespace(id=USER_ID, organization_id=ORG_ID, role=role)


def make_client(**kwargs) -> SimpleNamespace:
    defaults = {
        "id": CLIENT_ID,
        "full_name": "Иванов Иван Иванович",
        "phone": "+7 928 000-00-00",
        "contract_date": date(2025, 3, 15),
        "debt_amount": Decimal("350000.00"),
        "status": ClientStatus.ACTIVE,
        "assigned_manager_id": MANAGER_ID,
        "organization_id": ORG_ID,
        "is_deleted": False,
        "created_at": datetime(2025, 1, 1, 12, 0, 0),
        "updated_at": datetime(2025, 1, 1, 12, 0, 0),
    }
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


class TestExcelExport:
    def test_build_clients_workbook_for_owner(self, monkeypatch):
        client = make_client()
        manager = SimpleNamespace(id=MANAGER_ID, full_name="Менеджер Тест")
        db = MagicMock()
        db.scalars.return_value = [manager]
        monkeypatch.setattr(
            "app.services.excel_export.client_has_overdue_payments",
            lambda *_args, **_kwargs: True,
        )
        monkeypatch.setattr(
            "app.services.excel_export.client_overdue_amount",
            lambda *_args, **_kwargs: Decimal("5000.00"),
        )

        workbook = build_clients_workbook(db, make_user(), [client])
        content = workbook_to_bytes(workbook)
        loaded = load_workbook(BytesIO(content))

        assert loaded.sheetnames == ["Клиенты"]
        rows = list(loaded["Клиенты"].iter_rows(values_only=True))
        assert rows[0][0] == "ФИО"
        assert rows[1][0] == client.full_name
        assert rows[1][6] == "Да"
        assert rows[1][7] == 5000.0

    def test_build_clients_workbook_for_call_center(self):
        client = make_client()
        db = MagicMock()
        workbook = build_clients_workbook(db, make_user(UserRole.CALL_CENTER), [client])
        loaded = load_workbook(BytesIO(workbook_to_bytes(workbook)))
        rows = list(loaded["Клиенты"].iter_rows(values_only=True))

        assert len(rows[0]) == 4
        assert rows[1][0] == client.full_name

    def test_build_overdue_clients_workbook(self, monkeypatch):
        client = make_client()
        manager = SimpleNamespace(id=MANAGER_ID, full_name="Менеджер Тест")
        db = MagicMock()
        db.scalars.return_value = [manager]
        monkeypatch.setattr(
            "app.services.excel_export.client_overdue_amount",
            lambda *_args, **_kwargs: Decimal("12000.00"),
        )

        workbook = build_overdue_clients_workbook(db, [client])
        loaded = load_workbook(BytesIO(workbook_to_bytes(workbook)))
        rows = list(loaded["Просрочки"].iter_rows(values_only=True))

        assert rows[1][-1] == 12000.0

    def test_build_client_detail_workbook(self):
        detail = ClientDetailResponse(
            id=CLIENT_ID,
            organization_id=ORG_ID,
            assigned_manager_id=MANAGER_ID,
            full_name="Иванов Иван Иванович",
            phone="+7 928 000-00-00",
            contract_date=date(2025, 3, 15),
            debt_amount=Decimal("350000.00"),
            status=ClientStatus.ACTIVE,
            engagement_stage=EngagementStage.BANKRUPTCY,
            procedure_stage=ProcedureStage.CONTRACT_SIGNED,
            is_deleted=False,
            created_at=datetime(2025, 1, 1, 12, 0, 0),
            updated_at=datetime(2025, 1, 1, 12, 0, 0),
            has_overdue=True,
            installment_plan=InstallmentPlanResponse(
                id=uuid.uuid4(),
                client_id=CLIENT_ID,
                pricing_tier_id=uuid.uuid4(),
                total_amount=Decimal("400000.00"),
                start_date=date(2025, 3, 15),
                total_months=12,
                created_at=datetime(2025, 1, 1, 12, 0, 0),
            ),
            payment_schedule=[
                PaymentScheduleResponse(
                    id=uuid.uuid4(),
                    installment_plan_id=uuid.uuid4(),
                    month_number=1,
                    due_date=date(2025, 4, 15),
                    planned_amount=Decimal("10000.00"),
                    paid_amount=Decimal("3000.00"),
                    paid_date=None,
                    status=PaymentScheduleStatus.PARTIAL,
                )
            ],
            payments=[
                PaymentResponse(
                    id=uuid.uuid4(),
                    client_id=CLIENT_ID,
                    payment_schedule_id=uuid.uuid4(),
                    amount=Decimal("3000.00"),
                    payment_date=date(2025, 4, 10),
                    comment="Первый платёж",
                    created_by=USER_ID,
                    is_deleted=False,
                    is_refund=False,
                    created_at=datetime(2025, 1, 1, 12, 0, 0),
                )
            ],
            mandatory_payments=[],
        )

        workbook = build_client_detail_workbook(MagicMock(), detail, manager_name="Менеджер")
        loaded = load_workbook(BytesIO(workbook_to_bytes(workbook)))

        assert loaded.sheetnames == ["Карточка", "График", "Платежи", "Обязательные"]
        schedule_rows = list(loaded["График"].iter_rows(values_only=True))
        assert schedule_rows[1][0] == 1
        assert schedule_rows[1][6] == "Частично"

    def test_client_overdue_amount_counts_partial_past_due(self, monkeypatch):
        schedule = SimpleNamespace(
            planned_amount=Decimal("10000.00"),
            paid_amount=Decimal("2000.00"),
            status=PaymentScheduleStatus.PARTIAL,
            due_date=date(2020, 1, 1),
            deferred_until=None,
        )
        db = MagicMock()
        db.scalars.return_value = [schedule]
        monkeypatch.setattr(
            "app.services.excel_export.effective_due_date",
            lambda item: item.due_date,
        )

        amount = client_overdue_amount(db, CLIENT_ID)
        assert amount == Decimal("8000.00")
