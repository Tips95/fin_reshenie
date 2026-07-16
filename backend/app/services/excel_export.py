from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.client import Client
from app.models.enums import (
    ClientStatus,
    MandatoryPaymentType,
    PaymentScheduleStatus,
    UserRole,
)
from app.models.installment_plan import InstallmentPlan
from app.models.payment_schedule import PaymentSchedule
from app.models.user import User
from app.schemas.client import ClientDetailResponse
from app.services.access import client_has_overdue_payments
from app.services.schedule_dates import effective_due_date, is_schedule_overdue, schedule_remainder

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")

CLIENT_STATUS_LABELS = {
    ClientStatus.ACTIVE: "Активен",
    ClientStatus.COMPLETED: "Завершён",
    ClientStatus.DEFAULTED: "Просрочен",
    ClientStatus.CANCELLED: "Отменён",
}

SCHEDULE_STATUS_LABELS = {
    PaymentScheduleStatus.PENDING: "Ожидает",
    PaymentScheduleStatus.PAID: "Оплачен",
    PaymentScheduleStatus.PARTIAL: "Частично",
    PaymentScheduleStatus.OVERDUE: "Просрочен",
}

MANDATORY_TYPE_LABELS = {
    MandatoryPaymentType.DEPOSIT: "Депозит",
    MandatoryPaymentType.FINANCIAL_MANAGEMENT: "Финансовое управление",
    MandatoryPaymentType.COURT_FEE: "Судебная пошлина",
}


def _format_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d.%m.%Y")


def _format_money(value: Decimal | int | float | None) -> float | str:
    if value is None:
        return ""
    return float(value)


def _write_table(ws, headers: list[str], rows: list[list]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def _manager_names(db: Session, clients: list[Client]) -> dict:
    manager_ids = {client.assigned_manager_id for client in clients if client.assigned_manager_id}
    if not manager_ids:
        return {}
    managers = list(db.scalars(select(User).where(User.id.in_(manager_ids))))
    return {manager.id: manager.full_name for manager in managers}


def client_overdue_amount(db: Session, client_id) -> Decimal:
    today = date.today()
    schedules = list(
        db.scalars(
            select(PaymentSchedule)
            .join(InstallmentPlan, InstallmentPlan.id == PaymentSchedule.installment_plan_id)
            .where(InstallmentPlan.client_id == client_id)
        )
    )
    total = Decimal("0.00")
    for item in schedules:
        remainder = schedule_remainder(item)
        if remainder <= Decimal("0.00"):
            continue
        if is_schedule_overdue(item, today):
            total += remainder
    return total


def build_clients_workbook(
    db: Session,
    user: User,
    clients: list[Client],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    managers = _manager_names(db, clients)
    brief_mode = user.role == UserRole.CALL_CENTER

    if brief_mode:
        headers = ["ФИО", "Телефон", "Дата договора", "Статус"]
        rows = [
            [
                client.full_name,
                client.phone,
                _format_date(client.contract_date),
                CLIENT_STATUS_LABELS.get(client.status, client.status.value),
            ]
            for client in clients
        ]
    else:
        headers = [
            "ФИО",
            "Телефон",
            "Дата договора",
            "Сумма долга",
            "Статус",
            "Менеджер",
            "Просрочка",
            "Сумма просрочки",
        ]
        rows = []
        for client in clients:
            has_overdue = client_has_overdue_payments(db, client.id)
            rows.append(
                [
                    client.full_name,
                    client.phone,
                    _format_date(client.contract_date),
                    _format_money(client.debt_amount),
                    CLIENT_STATUS_LABELS.get(client.status, client.status.value),
                    managers.get(client.assigned_manager_id, ""),
                    "Да" if has_overdue else "Нет",
                    _format_money(client_overdue_amount(db, client.id)) if has_overdue else 0,
                ]
            )

    _write_table(ws, headers, rows)
    return wb


def build_overdue_clients_workbook(db: Session, clients: list[Client]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Просрочки"

    managers = _manager_names(db, clients)
    headers = [
        "ФИО",
        "Телефон",
        "Дата договора",
        "Сумма долга",
        "Статус",
        "Менеджер",
        "Сумма просрочки",
    ]
    rows = []
    for client in clients:
        rows.append(
            [
                client.full_name,
                client.phone,
                _format_date(client.contract_date),
                _format_money(client.debt_amount),
                CLIENT_STATUS_LABELS.get(client.status, client.status.value),
                managers.get(client.assigned_manager_id, ""),
                _format_money(client_overdue_amount(db, client.id)),
            ]
        )

    _write_table(ws, headers, rows)
    return wb


def build_client_detail_workbook(
    db: Session,
    detail: ClientDetailResponse,
    manager_name: str | None = None,
) -> Workbook:
    wb = Workbook()
    info_ws = wb.active
    info_ws.title = "Карточка"

    info_rows = [
        ["ФИО", detail.full_name],
        ["Телефон", detail.phone],
        ["Дата договора", _format_date(detail.contract_date)],
        ["Сумма долга", _format_money(detail.debt_amount)],
        ["Статус", CLIENT_STATUS_LABELS.get(detail.status, detail.status.value)],
        ["Менеджер", manager_name or ""],
        ["Просрочка", "Да" if detail.has_overdue else "Нет"],
    ]
    if detail.installment_plan:
        info_rows.extend(
            [
                ["Тариф", _format_money(detail.matched_tier.total_cost) if detail.matched_tier else ""],
                ["Срок, мес.", detail.installment_plan.total_months],
                ["Старт рассрочки", _format_date(detail.installment_plan.start_date)],
            ]
        )
    for row in info_rows:
        info_ws.append(row)
    for col in range(1, 3):
        info_ws.column_dimensions[get_column_letter(col)].width = 24

    schedule_ws = wb.create_sheet("График")
    schedule_headers = [
        "№",
        "Плановая дата",
        "Эффективная дата",
        "План",
        "Оплачено",
        "Остаток",
        "Статус",
        "Отсрочка до",
        "Комментарий отсрочки",
    ]
    schedule_rows = []
    for item in detail.payment_schedule:
        remainder = item.planned_amount - item.paid_amount
        if remainder < Decimal("0.00"):
            remainder = Decimal("0.00")
        schedule_rows.append(
            [
                item.month_number,
                _format_date(item.due_date),
                _format_date(effective_due_date(item)),
                _format_money(item.planned_amount),
                _format_money(item.paid_amount),
                _format_money(remainder),
                SCHEDULE_STATUS_LABELS.get(item.status, item.status.value),
                _format_date(item.deferred_until),
                item.deferral_comment or "",
            ]
        )
    _write_table(schedule_ws, schedule_headers, schedule_rows)

    payments_ws = wb.create_sheet("Платежи")
    payment_headers = ["Дата", "Сумма", "Возврат", "Комментарий"]
    payment_rows = [
        [
            _format_date(payment.payment_date),
            _format_money(payment.amount),
            "Да" if payment.is_refund else "Нет",
            payment.comment or "",
        ]
        for payment in detail.payments
    ]
    _write_table(payments_ws, payment_headers, payment_rows)

    mandatory_ws = wb.create_sheet("Обязательные")
    mandatory_headers = ["Тип", "План", "Оплачено", "Остаток", "Статус", "Дата оплаты", "Комментарий"]
    mandatory_rows = []
    for item in detail.mandatory_payments:
        remainder = item.planned_amount - item.paid_amount
        if remainder < Decimal("0.00"):
            remainder = Decimal("0.00")
        mandatory_rows.append(
            [
                MANDATORY_TYPE_LABELS.get(item.payment_type, item.payment_type.value),
                _format_money(item.planned_amount),
                _format_money(item.paid_amount),
                _format_money(remainder),
                item.status.value,
                _format_date(item.paid_date),
                item.comment or "",
            ]
        )
    _write_table(mandatory_ws, mandatory_headers, mandatory_rows)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
